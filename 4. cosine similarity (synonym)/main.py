#!/usr/bin/python

from numpy import load
from numpy import zeros
from openpyxl import load_workbook
from preprocessing import preprocessing
from sklearn.metrics.pairwise import cosine_similarity
from model.SingularValueDecomposition import SingularValueDecomposition

#1 = question, 2 = answer key, 3 = answer, 4 = testing, 5 = corpus
source = 5
source_path = 'source/' + str(source) + '/'

#load file
term_list = load(source_path + 'term_list.npy')
term_list = term_list.tolist()

# prepare list synonym
workbook = load_workbook("synonym.xlsx")
worksheet = workbook.active

for i in range(2, 42):
	worksheet['C'+str(i)] = str(preprocessing(worksheet['A'+str(i)].value))
	worksheet['D'+str(i)] = str(preprocessing(worksheet['B'+str(i)].value))

workbook.save('synonym_stemmed.xlsx')

#get svd

svd_tf_idf = SingularValueDecomposition()
svd_widf = SingularValueDecomposition()
svd_midf = SingularValueDecomposition()

svd_tf_idf.setMatrixDecomposition(load(source_path + str(source)+'.svd.tf_idf.u.npy'),load(source_path + str(source)+'.svd.tf_idf.s.npy'),load(source_path + str(source)+'.svd.tf_idf.vt.npy'))
svd_widf.setMatrixDecomposition(load(source_path + str(source)+'.svd.widf.u.npy'),load(source_path + str(source)+'.svd.widf.s.npy'),load(source_path + str(source)+'.svd.widf.vt.npy'))
svd_midf.setMatrixDecomposition(load(source_path + str(source)+'.svd.midf.u.npy'),load(source_path + str(source)+'.svd.midf.s.npy'),load(source_path + str(source)+'.svd.midf.vt.npy'))
	
#eksperiment

k_dimension = 285

svd_tf_idf.setReductionSVD(k_dimension)
svd_widf.setReductionSVD(k_dimension)
svd_midf.setReductionSVD(k_dimension)

workbook = load_workbook("synonym_stemmed.xlsx")
worksheet = workbook.active

for i in range(2, 42):
	
	m1 = zeros([len(term_list), 1])	
	m2 = zeros([len(term_list), 1])	
	
	term = worksheet['A'+str(i)].value
	ref  = worksheet['B'+str(i)].value
	term_stemmed = worksheet['C'+str(i)].value
	ref_stemmed  = worksheet['D'+str(i)].value
	
	print term + " - " + ref
	
	score_tf_idf = 0
	score_widf = 0
	score_midf = 0
			
	if term_stemmed in term_list:
		m1[term_list.index(term_stemmed),0]=1
	
		if ref_stemmed in term_list:
			m2[term_list.index(ref_stemmed),0]=1

			v1 = m1[:,0:1]	
			v2 = m2[:,0:1]
			
			score = cosine_similarity(v1.transpose(), v2.transpose())[0][0]

			score_tf_idf = cosine_similarity(svd_tf_idf.createQuery(v1), svd_tf_idf.createQuery(v2))[0][0]
			score_widf = cosine_similarity(svd_widf.createQuery(v1), svd_widf.createQuery(v2))[0][0]
			score_midf = cosine_similarity(svd_midf.createQuery(v1), svd_midf.createQuery(v2))[0][0]
				
			worksheet['E'+str(i)] = score_tf_idf
			worksheet['F'+str(i)] = score_widf
			worksheet['G'+str(i)] = score_midf
					
	print "   cosine : " + str(score) 
	print "   tf_idf : " + str(score_tf_idf) 
	print "   widf   : " + str(score_widf)
	print "   midf   : " + str(score_midf)	
	
workbook.save('result.xlsx')
 